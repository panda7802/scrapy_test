# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://doc.scrapy.org/en/latest/topics/item-pipeline.html
# import xlwt
import logging

import openpyxl
from openpyxl import Workbook

from search_food.items import ShopItem, TypeItem
import os


class SearchFoodPipeline(object):
    def process_item(self, item, spider):
        return item


class ShopPipeline(object):

    def __init__(self):
        logging.debug("-------ShopPipeline __init__---------")
        self.file_full_dir = "./res_dir"
        self.file_full_path = self.file_full_dir + "/shops.xlsx"
        self.types = []
        if not os.path.exists(self.file_full_path):
            self.wb = Workbook()  # class实例化
            self.ws = self.wb.active  # 激活工作表
            if not os.path.exists(self.file_full_dir):
                os.makedirs(self.file_full_dir)
            self.wb.save(self.file_full_path)  # 保存文件
        else:
            self.wb = openpyxl.load_workbook(self.file_full_path)

    def init_ws(self, type_name):
        # type_index = self.types.index(type_name)
        if type_name in self.types:
            self.ws = self.wb.index(type_name)
        else:
            self.types.append(type_name)
            if self.types.__len__() > 1:
                self.ws = self.wb.create_sheet(type_name)
            else:
                self.ws = self.wb.active
                self.ws.title = type_name
            self.ws.append(['店名', '价格', '电话', '图片数', '好评', '中评', '差评', '好评率', '链接'])  # 加入一行数据

    def open_spider(self, spider):
        logging.debug("-------ShopPipeline open_spider---------")

    def add_line(self, item):
        """
        增加行
        :param item:
        :return:
        """
        line = [item['name'], item['price'], item['phone'], int(item['pic']), int(item['good']), int(item['common']),
                int(item['bad']), -1, item['url']]
        type_name = item['type_name']
        if type_name not in self.types:
            self.init_ws(type_name)
        self.ws = self.wb.get_sheet_by_name(type_name)
        self.ws.append(line)
        name_cell = "A%d" % self.ws.max_row
        self.ws[name_cell].hyperlink = item['url']
        self.wb.save(self.file_full_path)  # 保存文件

    def upd_line(self, item):
        """
        修改行
        :param item:
        :return:
        """
        row_num = int(item['row_num']) + 1
        all_commons = int(item['good']) + int(item['common']) + int(item['bad'])
        hpl = -1
        if all_commons != 0:
            hpl = float(item['good']) / all_commons
        type_name = item['type_name']
        self.ws = self.wb.get_sheet_by_name(type_name)
        self.ws["B%d" % row_num] = item['price']
        self.ws["C%d" % row_num] = item['phone']
        self.ws["D%d" % row_num] = item['pic']
        self.ws["E%d" % row_num] = item['good']
        self.ws["F%d" % row_num] = item['common']
        self.ws["G%d" % row_num] = item['bad']
        self.ws["H%d" % row_num] = hpl
        self.wb.save(self.file_full_path)  # 保存文件

    def handler_shop(self, item):
        """
        处理商店信息
        :param item:
        :return:
        """
        print(item)
        if int(item['row_num']) >= 0:
            self.upd_line(item)
        else:
            self.add_line(item)

    def handler_type(self, item):
        """
        处理类型信息
        :param item:
        :return:
        """
        wb_index = 0
        # logging.debug("item : " + item['name'])
        self.init_ws(item['name'])
        # self.ws = self.wb.index(wb_index)

    def process_item(self, item, spider):
        # logging.debug(type(item))
        # logging.debug(type(item) is ShopItem)
        if type(item) is ShopItem:
            self.handler_shop(item)
        if type(item) is TypeItem:
            self.handler_type(item)
        return item

    def close_spider(self, spider):
        self.wb.save(self.file_full_path)
