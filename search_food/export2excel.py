import datetime
import os
import sys
import time

import django
import openpyxl
from django.db.models import Q
from openpyxl import Workbook

DJANGO_PROJECT_PATH = '../../lxdzx_server'
DJANGO_SETTINGS_MODULE = 'lxdzx_server.settings'
sys.path.insert(0, DJANGO_PROJECT_PATH)
os.environ['DJANGO_SETTINGS_MODULE'] = DJANGO_SETTINGS_MODULE
# os.environ.setdefault("DJANGO_SETTINGS_MODULE", DJANGO_SETTINGS_MODULE    )
# application = django.core.handlers.wsgi.WSGIHandler()
print("===========setting over===========")
django.setup()

from dzdp.models import DzdpType, DzdpShop, DzdpShopType

file_full_dir = "./res_dir"
file_full_path = file_full_dir + "/shops.xlsx"
file_full_path_bak = file_full_dir + "/shops_bak.xlsx"


def export_types(wb):
    """
    导出类型
    :return:
    """
    types = DzdpType.objects.filter(parent_type_id=1).all()
    for item in types:
        ws = wb.create_sheet(item.name)
        ws.append(['店名', '价格', '电话', '图片数', '好评', '中评', '差评', '好评率', '链接'])  # 加入一行数据


def export_shops_by_type(wb, type_name):
    """
    导出商店
    :return:
    """
    # shops = DzdpShop.objects.filter(type__name=type_name).all()
    sql = 'select a.* from v_dzdpshop a, dzdp_dzdpshoptype b,dzdp_dzdptype c ' \
          ' where c.name = \'%s\' and a.id = b.shop_id  and b.type_id = c.id ' % type_name
    shops = DzdpShop.objects.raw(sql)
    count = len(shops)

    #修改头和标签
    wb_type = Workbook()
    ws = wb_type.active
    ws.title = type_name
    ws.append(['店名', '价格', '电话', '图片数', '好评', '中评', '差评', '好评率', '链接'])  # 加入一行数据

    # ws = wb.get_sheet_by_name(type_name)
    for index, shop in enumerate(shops):
        # if index > 10:
        #     break
        all_commons = shop.good + shop.common + shop.bad
        hpl = -1
        if all_commons != 0:
            hpl = float(shop.good) / all_commons
        line = [shop.name, shop.price, shop.phone, shop.pic, shop.good, shop.common,
                shop.bad, hpl, shop.url]
        ws.append(line)
        name_cell = "A%d" % ws.max_row
        ws[name_cell].hyperlink = shop.url
        print("%s -- %s , %s : %d/%d" % (time.strftime("%H:%M.%S"), type_name, shop.name, index + 1, count))
        if index % 100 == 0:
            # wb.save(file_full_path)  # 保存文件
            wb_type.save("%s/%s.xlsx" % (file_full_dir, type_name))

    wb_type.save(file_full_path)  # 保存文件


if __name__ == '__main__':
    print("-------start export---------")
    # obj_file = open(file_full_path)
    # if obj_file is not None:
    #     delete()
    if not os.path.exists(file_full_path):
        if not os.path.exists(file_full_dir):
            os.makedirs(file_full_dir)

    if os.path.exists(file_full_path):
        os.remove(file_full_path)

    wb = Workbook()  # class实例化

    # 导出类型
    export_types(wb)

    def_sheet = wb.get_sheet_by_name('Sheet')
    wb.remove(def_sheet)
    wb.save(file_full_path)  # 保存文件

    # 根据类型保持内容
    types = DzdpType.objects.filter(parent_type_id=1).all()
    for index, item in enumerate(types):
        # if index > 2:
        #     break
        export_shops_by_type(wb, item.name)
        wb.save(file_full_path)
